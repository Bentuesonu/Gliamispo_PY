import openpyxl, io
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)


def _style_header_row(ws, row_idx, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws):
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[
            get_column_letter(col[0].column)
        ].width = min(max_len + 4, 50)


def export_breakdown(db, project_id: int) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Breakdown"
    headers = ["#", "Location", "INT/EXT", "G/N", "Pagine", "Categoria", "Elemento"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    rows = db.execute('''
        SELECT s.scene_number, s.location, s.int_ext, s.day_night,
               s.page_start_whole, s.page_start_eighths,
               se.category, se.element_name
        FROM scenes s
        LEFT JOIN scene_elements se ON se.scene_id = s.id
        WHERE s.project_id = ?
        ORDER BY s.scene_number, se.category, se.element_name
    ''', (project_id,)).fetchall()
    for r in rows:
        pages = f'{r[4]} {r[5]}/8' if r[5] else str(r[4])
        ws.append([r[0], r[1], r[2], r[3], pages, r[6], r[7]])
    _autofit(ws)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def export_budget(db, project_id: int) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Budget"
    headers = ["Codice", "Voce", "Qtà", "Tariffa", "Totale", "Note"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    rows = db.execute('''
        SELECT ba.account_code, ba.account_name, bd.units, bd.rate,
               bd.units * bd.rate, bd.description
        FROM budget_accounts ba
        JOIN budget_details bd ON bd.account_id = ba.id
        WHERE ba.project_id = ?
        ORDER BY ba.account_code, ba.account_name
    ''', (project_id,)).fetchall()
    for r in rows:
        ws.append(list(r))
    _autofit(ws)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def export_stripboard(db, project_id: int) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Stripboard"
    headers = ["Giorno", "Scena", "Location", "INT/EXT", "G/N", "Pagine", "Bloccata"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    rows = db.execute('''
        SELECT se.shooting_day, s.scene_number, s.location, s.int_ext, s.day_night,
               s.page_start_whole, s.page_start_eighths,
               s.page_end_whole, s.page_end_eighths, s.is_locked
        FROM schedule_entries se
        JOIN scenes s ON se.scene_id = s.id
        WHERE se.project_id = ?
        ORDER BY se.shooting_day, se.position
    ''', (project_id,)).fetchall()
    for r in rows:
        pages_start = f'{r[5]} {r[6]}/8' if r[6] else str(r[5] or '')
        pages_end = f'{r[7]} {r[8]}/8' if r[8] else str(r[7] or '')
        pages = f'{pages_start} - {pages_end}' if pages_end else pages_start
        ws.append([r[0], r[1], r[2], r[3], r[4], pages, '🔒' if r[9] else ''])
    _autofit(ws)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def export_oneliner(db, project_id: int) -> bytes:
    """Export One-Liner schedule to Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "One-Liner"
    headers = ["Giorno", "Scena", "INT/EXT", "G/N", "Location", "Pagine", "Cast", "Sinossi"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    try:
        entries = db.execute('''
            SELECT se.shooting_day, s.scene_number, s.int_ext, s.day_night,
                   s.location, s.page_start_whole, s.page_start_eighths,
                   s.page_end_whole, s.page_end_eighths, s.id, s.synopsis
            FROM schedule_entries se
            JOIN scenes s ON se.scene_id = s.id
            WHERE se.project_id = ?
            ORDER BY se.shooting_day, se.position
        ''', (project_id,)).fetchall()
    except Exception:
        entries = []

    if not entries:
        entries = db.execute('''
            SELECT 1, scene_number, int_ext, day_night, location,
                   page_start_whole, page_start_eighths,
                   page_end_whole, page_end_eighths, id, synopsis
            FROM scenes WHERE project_id = ? ORDER BY id
        ''', (project_id,)).fetchall()

    for e in entries:
        # Get cast for this scene
        scene_id = e[9]
        cast_rows = db.execute(
            "SELECT element_name FROM scene_elements"
            " WHERE scene_id = ? AND category = 'Cast'"
            " ORDER BY element_name",
            (scene_id,)
        ).fetchall()
        cast = ', '.join(c[0] for c in cast_rows)

        pages_start = f'{e[5]} {e[6]}/8' if e[6] else str(e[5] or '')
        pages_end = f'{e[7]} {e[8]}/8' if e[8] else str(e[7] or '')
        pages = f'{pages_start} - {pages_end}' if pages_end else pages_start
        synopsis = e[10] or ''

        ws.append([e[0], e[1], e[2], e[3], e[4], pages, cast, synopsis[:100]])

    _autofit(ws)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def export_dood(db, project_id: int) -> bytes:
    """Export Day Out of Days matrix to Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Day Out of Days"

    # Leggi attori
    cast = db.execute(
        "SELECT DISTINCT element_name FROM scene_elements"
        " WHERE category = 'Cast' AND scene_id IN"
        " (SELECT id FROM scenes WHERE project_id = ?)"
        " ORDER BY element_name",
        (project_id,)
    ).fetchall()
    cast_names = [c[0] for c in cast]

    # Leggi giorni
    try:
        days = db.execute(
            "SELECT DISTINCT shooting_day FROM schedule_entries"
            " WHERE project_id = ? ORDER BY shooting_day",
            (project_id,)
        ).fetchall()
        day_nums = [d[0] for d in days]
    except Exception:
        day_nums = []

    if not day_nums:
        n = db.execute(
            "SELECT COUNT(*) FROM scenes WHERE project_id = ?",
            (project_id,)
        ).fetchone()[0]
        day_nums = list(range(1, n + 1))

    # Mappa scena → giorni
    scene_days = {}
    try:
        entries = db.execute(
            "SELECT scene_id, shooting_day FROM schedule_entries WHERE project_id = ?",
            (project_id,)
        ).fetchall()
        for e in entries:
            scene_days.setdefault(e[0], []).append(e[1])
    except Exception:
        pass

    # Calcola matrice
    matrix = {}
    for actor in cast_names:
        sc_rows = db.execute(
            "SELECT scene_id FROM scene_elements"
            " WHERE element_name = ? AND category = 'Cast'"
            " AND scene_id IN (SELECT id FROM scenes WHERE project_id = ?)",
            (actor, project_id),
        ).fetchall()
        work = set()
        for sf in sc_rows:
            for d in scene_days.get(sf[0], []):
                work.add(d)
        if work:
            sorted_work = sorted(work)
            first_day, last_day = sorted_work[0], sorted_work[-1]
            actor_data = {}
            for d in day_nums:
                if d in work:
                    actor_data[d] = 'W'
                elif first_day <= d <= last_day:
                    actor_data[d] = 'H'
            matrix[actor] = actor_data

    # Scrivi header
    headers = ["Attore"] + [f"G{d}" for d in day_nums]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    # Scrivi righe
    STATUS_COLORS = {
        'W': PatternFill("solid", fgColor="1E8449"),  # Verde - Working
        'H': PatternFill("solid", fgColor="707B7C"),  # Grigio - Hold
        'T': PatternFill("solid", fgColor="B7950B"),  # Giallo - Travel
        'F': PatternFill("solid", fgColor="922B21"),  # Rosso - Finished
    }
    for actor in cast_names:
        row = [actor]
        for d in day_nums:
            status = matrix.get(actor, {}).get(d, '')
            row.append(status)
        ws.append(row)
        # Applica colori
        for col_idx, d in enumerate(day_nums, start=2):
            status = matrix.get(actor, {}).get(d, '')
            if status in STATUS_COLORS:
                ws.cell(ws.max_row, col_idx).fill = STATUS_COLORS[status]
                ws.cell(ws.max_row, col_idx).font = Font(color="FFFFFF", bold=True)
                ws.cell(ws.max_row, col_idx).alignment = Alignment(horizontal="center")

    _autofit(ws)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()
