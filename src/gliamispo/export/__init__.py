"""
Export module - supporta sia Excel (.xlsx) che PDF.

Uso:
    from gliamispo.export import export_breakdown, export_oneliner, Format

    # Excel (default)
    data = export_breakdown(db, project_id)
    data = export_breakdown(db, project_id, fmt=Format.EXCEL)

    # PDF
    data = export_breakdown(db, project_id, fmt=Format.PDF)
"""

from enum import Enum, auto

# Excel exports
from .excel_export import (
    export_breakdown as _excel_breakdown,
    export_budget as _excel_budget,
    export_stripboard as _excel_stripboard,
    export_oneliner as _excel_oneliner,
    export_dood as _excel_dood,
)

# PDF exports
from .pdf_exporter import (
    BasePDFExporter,
    OneLinerExporter,
    DayOutOfDaysExporter,
    HAS_FPDF,
    _sanitize_text,
)
from .call_sheet_pdf import (
    CallSheetGenerator,
    CallSheetPDF,
    generate_call_sheet_pdf,
)


class Format(Enum):
    """Formato di export."""
    EXCEL = auto()
    PDF = auto()


def export_breakdown(db, project_id: int, fmt: Format = Format.EXCEL) -> bytes:
    """
    Esporta breakdown scene.

    Args:
        db: connessione database
        project_id: ID progetto
        fmt: Format.EXCEL o Format.PDF

    Returns:
        bytes del file generato
    """
    if fmt == Format.EXCEL:
        return _excel_breakdown(db, project_id)
    else:
        return _export_breakdown_pdf(db, project_id)


def export_budget(db, project_id: int, fmt: Format = Format.EXCEL) -> bytes:
    """
    Esporta budget.

    Args:
        db: connessione database
        project_id: ID progetto
        fmt: Format.EXCEL o Format.PDF

    Returns:
        bytes del file generato
    """
    if fmt == Format.EXCEL:
        return _excel_budget(db, project_id)
    else:
        return _export_budget_pdf(db, project_id)


def export_stripboard(db, project_id: int, fmt: Format = Format.EXCEL) -> bytes:
    """
    Esporta stripboard.

    Args:
        db: connessione database
        project_id: ID progetto
        fmt: Format.EXCEL o Format.PDF

    Returns:
        bytes del file generato
    """
    if fmt == Format.EXCEL:
        return _excel_stripboard(db, project_id)
    else:
        return _export_stripboard_pdf(db, project_id)


def export_oneliner(db, project_id: int, fmt: Format = Format.EXCEL) -> bytes:
    """
    Esporta One-Liner schedule.

    Args:
        db: connessione database
        project_id: ID progetto
        fmt: Format.EXCEL o Format.PDF

    Returns:
        bytes del file generato
    """
    if fmt == Format.EXCEL:
        return _excel_oneliner(db, project_id)
    else:
        return _export_oneliner_pdf(db, project_id)


def export_dood(db, project_id: int, fmt: Format = Format.EXCEL) -> bytes:
    """
    Esporta Day Out of Days.

    Args:
        db: connessione database
        project_id: ID progetto
        fmt: Format.EXCEL o Format.PDF

    Returns:
        bytes del file generato
    """
    if fmt == Format.EXCEL:
        return _excel_dood(db, project_id)
    else:
        return _export_dood_pdf(db, project_id)


def export_call_sheet(db, call_sheet_id: int, fmt: Format = Format.PDF) -> bytes:
    """
    Esporta foglio di lavorazione (call sheet).

    Args:
        db: connessione database
        call_sheet_id: ID del call sheet
        fmt: Format.PDF (default) o Format.EXCEL

    Returns:
        bytes del file generato
    """
    if fmt == Format.PDF:
        return generate_call_sheet_pdf(db, call_sheet_id)
    else:
        return _export_call_sheet_excel(db, call_sheet_id)


# ---------------------------------------------------------------------------
# Implementazioni PDF per breakdown, budget, stripboard
# ---------------------------------------------------------------------------

def _export_breakdown_pdf(db, project_id: int) -> bytes:
    """Esporta breakdown in PDF."""
    if not HAS_FPDF:
        return b""
    from fpdf import FPDF

    rows = db.execute('''
        SELECT s.scene_number, s.location, s.int_ext, s.day_night,
               s.page_start_whole, s.page_start_eighths,
               se.category, se.element_name
        FROM scenes s
        LEFT JOIN scene_elements se ON se.scene_id = s.id
        WHERE s.project_id = ?
        ORDER BY s.scene_number, se.category, se.element_name
    ''', (project_id,)).fetchall()

    project = db.execute(
        "SELECT title FROM projects WHERE id = ?", (project_id,)
    ).fetchone()
    title = _sanitize_text(project[0] if project else "Progetto")

    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_margins(10, 15, 10)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Header
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_text_color(10, 36, 82)
    pdf.cell(0, 8, 'GLIAMISPO', ln=True)
    pdf.set_font('Helvetica', '', 10)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(0, 6, f'{title}  -  Breakdown', ln=True)
    pdf.ln(4)
    pdf.set_draw_color(200, 200, 200)
    pdf.line(10, pdf.get_y(), pdf.w - 10, pdf.get_y())
    pdf.ln(4)

    # Colonne
    cols = [('#', 12), ('Location', 50), ('I/E', 15), ('G/N', 15),
            ('Pag.', 18), ('Categoria', 40), ('Elemento', 70)]

    pdf.set_fill_color(26, 82, 118)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 8)
    for name, w in cols:
        pdf.cell(w, 6, name, border=0, fill=True, align='C')
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Helvetica', '', 7)
    for i, r in enumerate(rows):
        fill = i % 2 == 0
        pdf.set_fill_color(244, 246, 246) if fill else pdf.set_fill_color(255, 255, 255)
        pages = f'{r[4]} {r[5]}/8' if r[5] else str(r[4] or '')

        pdf.cell(cols[0][1], 5, _sanitize_text(str(r[0] or '')), fill=fill, align='C')
        pdf.cell(cols[1][1], 5, _sanitize_text(str(r[1] or ''))[:30], fill=fill)
        pdf.cell(cols[2][1], 5, _sanitize_text(str(r[2] or '')), fill=fill, align='C')
        pdf.cell(cols[3][1], 5, _sanitize_text(str(r[3] or '')), fill=fill, align='C')
        pdf.cell(cols[4][1], 5, pages, fill=fill, align='C')
        pdf.cell(cols[5][1], 5, _sanitize_text(str(r[6] or ''))[:25], fill=fill)
        pdf.cell(cols[6][1], 5, _sanitize_text(str(r[7] or ''))[:45], fill=fill)
        pdf.ln()

    return bytes(pdf.output())


def _export_budget_pdf(db, project_id: int) -> bytes:
    """Esporta budget in PDF."""
    if not HAS_FPDF:
        return b""
    from fpdf import FPDF

    rows = db.execute('''
        SELECT ba.account_code, ba.account_name, bd.units, bd.rate,
               bd.units * bd.rate, bd.description
        FROM budget_accounts ba
        JOIN budget_details bd ON bd.account_id = ba.id
        WHERE ba.project_id = ?
        ORDER BY ba.account_code, ba.account_name
    ''', (project_id,)).fetchall()

    project = db.execute(
        "SELECT title FROM projects WHERE id = ?", (project_id,)
    ).fetchone()
    title = _sanitize_text(project[0] if project else "Progetto")

    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.set_margins(15, 15, 15)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Header
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_text_color(10, 36, 82)
    pdf.cell(0, 8, 'GLIAMISPO', ln=True)
    pdf.set_font('Helvetica', '', 10)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(0, 6, f'{title}  -  Budget', ln=True)
    pdf.ln(4)
    pdf.set_draw_color(200, 200, 200)
    pdf.line(15, pdf.get_y(), pdf.w - 15, pdf.get_y())
    pdf.ln(4)

    cols = [('Codice', 25), ('Voce', 50), ('Qtà', 20), ('Tariffa', 25),
            ('Totale', 25), ('Note', 35)]

    pdf.set_fill_color(26, 82, 118)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 8)
    for name, w in cols:
        pdf.cell(w, 6, name, border=0, fill=True, align='C')
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Helvetica', '', 7)
    total_budget = 0
    for i, r in enumerate(rows):
        fill = i % 2 == 0
        pdf.set_fill_color(244, 246, 246) if fill else pdf.set_fill_color(255, 255, 255)
        total_budget += r[4] or 0

        pdf.cell(cols[0][1], 5, _sanitize_text(str(r[0] or '')), fill=fill, align='C')
        pdf.cell(cols[1][1], 5, _sanitize_text(str(r[1] or ''))[:30], fill=fill)
        pdf.cell(cols[2][1], 5, _sanitize_text(str(r[2] or '')), fill=fill, align='R')
        pdf.cell(cols[3][1], 5, f'EUR {r[3]:.2f}' if r[3] else '', fill=fill, align='R')
        pdf.cell(cols[4][1], 5, f'EUR {r[4]:.2f}' if r[4] else '', fill=fill, align='R')
        pdf.cell(cols[5][1], 5, _sanitize_text(str(r[5] or ''))[:20], fill=fill)
        pdf.ln()

    # Totale
    pdf.ln(2)
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(sum(c[1] for c in cols[:4]), 6, 'TOTALE', align='R')
    pdf.cell(cols[4][1], 6, f'EUR {total_budget:.2f}', align='R')
    pdf.ln()

    return bytes(pdf.output())


def _export_stripboard_pdf(db, project_id: int) -> bytes:
    """Esporta stripboard in PDF."""
    if not HAS_FPDF:
        return b""
    from fpdf import FPDF

    rows = db.execute('''
        SELECT se.shooting_day, s.scene_number, s.location, s.int_ext, s.day_night,
               s.page_start_whole, s.page_start_eighths,
               s.page_end_whole, s.page_end_eighths, s.is_locked
        FROM schedule_entries se
        JOIN scenes s ON se.scene_id = s.id
        WHERE se.project_id = ?
        ORDER BY se.shooting_day, se.position
    ''', (project_id,)).fetchall()

    project = db.execute(
        "SELECT title FROM projects WHERE id = ?", (project_id,)
    ).fetchone()
    title = _sanitize_text(project[0] if project else "Progetto")

    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_margins(10, 15, 10)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Header
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_text_color(10, 36, 82)
    pdf.cell(0, 8, 'GLIAMISPO', ln=True)
    pdf.set_font('Helvetica', '', 10)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(0, 6, f'{title}  -  Stripboard', ln=True)
    pdf.ln(4)
    pdf.set_draw_color(200, 200, 200)
    pdf.line(10, pdf.get_y(), pdf.w - 10, pdf.get_y())
    pdf.ln(4)

    cols = [('Giorno', 18), ('Scena', 18), ('Location', 80), ('I/E', 15),
            ('G/N', 15), ('Pagine', 25), ('Lock', 12)]

    pdf.set_fill_color(26, 82, 118)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 8)
    for name, w in cols:
        pdf.cell(w, 6, name, border=0, fill=True, align='C')
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Helvetica', '', 7)
    for i, r in enumerate(rows):
        fill = i % 2 == 0
        pdf.set_fill_color(244, 246, 246) if fill else pdf.set_fill_color(255, 255, 255)

        pages_start = f'{r[5]} {r[6]}/8' if r[6] else str(r[5] or '')
        pages_end = f'{r[7]} {r[8]}/8' if r[8] else str(r[7] or '')
        pages = f'{pages_start} - {pages_end}' if pages_end else pages_start
        locked = 'X' if r[9] else ''

        pdf.cell(cols[0][1], 5, _sanitize_text(str(r[0] or '')), fill=fill, align='C')
        pdf.cell(cols[1][1], 5, _sanitize_text(str(r[1] or '')), fill=fill, align='C')
        pdf.cell(cols[2][1], 5, _sanitize_text(str(r[2] or ''))[:45], fill=fill)
        pdf.cell(cols[3][1], 5, _sanitize_text(str(r[3] or '')), fill=fill, align='C')
        pdf.cell(cols[4][1], 5, _sanitize_text(str(r[4] or '')), fill=fill, align='C')
        pdf.cell(cols[5][1], 5, pages, fill=fill, align='C')
        pdf.cell(cols[6][1], 5, locked, fill=fill, align='C')
        pdf.ln()

    return bytes(pdf.output())


def _export_oneliner_pdf(db, project_id: int) -> bytes:
    """Esporta One-Liner in PDF usando OneLinerExporter."""
    try:
        entries = db.execute('''
            SELECT se.shooting_day, s.scene_number, s.int_ext, s.day_night,
                   s.location, s.page_start_whole, s.page_start_eighths,
                   s.page_end_whole, s.page_end_eighths, s.id, s.synopsis
            FROM schedule_entries se
            JOIN scenes s ON se.scene_id = s.id
            WHERE se.project_id = ?
            ORDER BY se.shooting_day, se.position
        ''', (project_id,)).fetchall()
    except Exception:
        entries = []

    if not entries:
        entries = db.execute('''
            SELECT 1, scene_number, int_ext, day_night, location,
                   page_start_whole, page_start_eighths,
                   page_end_whole, page_end_eighths, id, synopsis
            FROM scenes WHERE project_id = ? ORDER BY id
        ''', (project_id,)).fetchall()

    rows = []
    for e in entries:
        scene_id = e[9]
        cast_rows = db.execute(
            "SELECT element_name FROM scene_elements"
            " WHERE scene_id = ? AND category = 'Cast'"
            " ORDER BY element_name",
            (scene_id,)
        ).fetchall()
        cast = ', '.join(c[0] for c in cast_rows)

        pages_start = f'{e[5]} {e[6]}/8' if e[6] else str(e[5] or '')
        pages_end = f'{e[7]} {e[8]}/8' if e[8] else str(e[7] or '')
        pages = f'{pages_start} - {pages_end}' if pages_end else pages_start
        synopsis = e[10] or ''

        rows.append((e[0], e[1], e[2], e[3], e[4], pages, cast, synopsis[:100]))

    project = db.execute(
        "SELECT title FROM projects WHERE id = ?", (project_id,)
    ).fetchone()
    title = project[0] if project else "Progetto"

    exporter = OneLinerExporter(title)
    result = exporter.export(rows)
    return result if result else b""


def _export_dood_pdf(db, project_id: int) -> bytes:
    """Esporta Day Out of Days in PDF usando DayOutOfDaysExporter."""
    # Leggi attori
    cast = db.execute(
        "SELECT DISTINCT element_name FROM scene_elements"
        " WHERE category = 'Cast' AND scene_id IN"
        " (SELECT id FROM scenes WHERE project_id = ?)"
        " ORDER BY element_name",
        (project_id,)
    ).fetchall()
    cast_names = [c[0] for c in cast]

    # Leggi giorni
    try:
        days = db.execute(
            "SELECT DISTINCT shooting_day FROM schedule_entries"
            " WHERE project_id = ? ORDER BY shooting_day",
            (project_id,)
        ).fetchall()
        day_nums = [d[0] for d in days]
    except Exception:
        day_nums = []

    if not day_nums:
        n = db.execute(
            "SELECT COUNT(*) FROM scenes WHERE project_id = ?",
            (project_id,)
        ).fetchone()[0]
        day_nums = list(range(1, n + 1))

    # Mappa scena → giorni
    scene_days = {}
    try:
        entries = db.execute(
            "SELECT scene_id, shooting_day FROM schedule_entries WHERE project_id = ?",
            (project_id,)
        ).fetchall()
        for e in entries:
            scene_days.setdefault(e[0], []).append(e[1])
    except Exception:
        pass

    # Calcola matrice
    matrix = {}
    for actor in cast_names:
        sc_rows = db.execute(
            "SELECT scene_id FROM scene_elements"
            " WHERE element_name = ? AND category = 'Cast'"
            " AND scene_id IN (SELECT id FROM scenes WHERE project_id = ?)",
            (actor, project_id),
        ).fetchall()
        work = set()
        for sf in sc_rows:
            for d in scene_days.get(sf[0], []):
                work.add(d)
        if work:
            sorted_work = sorted(work)
            first_day, last_day = sorted_work[0], sorted_work[-1]
            actor_data = {}
            for d in day_nums:
                if d in work:
                    actor_data[d] = 'W'
                elif first_day <= d <= last_day:
                    actor_data[d] = 'H'
            matrix[actor] = actor_data

    project = db.execute(
        "SELECT title FROM projects WHERE id = ?", (project_id,)
    ).fetchone()
    title = project[0] if project else "Progetto"

    exporter = DayOutOfDaysExporter(title)
    result = exporter.export(cast_names, day_nums, matrix)
    return result if result else b""


def _export_call_sheet_excel(db, call_sheet_id: int) -> bytes:
    """Esporta call sheet in Excel."""
    import openpyxl
    import io
    from openpyxl.styles import PatternFill, Font, Alignment

    cs = db.execute(
        """
        SELECT cs.*, sd.day_number, sd.shoot_date, sd.location_primary
        FROM call_sheets cs
        JOIN shooting_days sd ON sd.id = cs.shooting_day_id
        WHERE cs.id = ?
        """,
        (call_sheet_id,),
    ).fetchone()

    if cs is None:
        return b""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Call Sheet"

    HEADER_FILL = PatternFill("solid", fgColor="1A5276")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)

    # Info generali
    ws.append(["FOGLIO DI LAVORAZIONE"])
    ws.merge_cells('A1:F1')
    ws['A1'].font = Font(bold=True, size=14)

    ws.append([])
    ws.append(["Giorno:", cs[4] if len(cs) > 4 else "", "Data:", cs[5] if len(cs) > 5 else ""])
    ws.append(["Location:", cs[6] if len(cs) > 6 else ""])
    ws.append(["Chiamata:", cs[1] if len(cs) > 1 else "07:00"])
    ws.append([])

    # Cast
    ws.append(["CAST"])
    ws['A7'].fill = HEADER_FILL
    ws['A7'].font = HEADER_FONT

    cast = db.execute(
        "SELECT actor_name, character_name, call_time"
        " FROM call_sheet_cast WHERE call_sheet_id = ? ORDER BY call_time",
        (call_sheet_id,),
    ).fetchall()

    ws.append(["Attore", "Personaggio", "Chiamata"])
    for c in cast:
        ws.append(list(c))

    ws.append([])

    # Crew
    ws.append(["TROUPE"])
    row_idx = ws.max_row
    ws.cell(row=row_idx, column=1).fill = HEADER_FILL
    ws.cell(row=row_idx, column=1).font = HEADER_FONT

    crew = db.execute(
        "SELECT crew_member_name, department, call_time"
        " FROM call_sheet_crew WHERE call_sheet_id = ? ORDER BY department, call_time",
        (call_sheet_id,),
    ).fetchall()

    ws.append(["Nome", "Dipartimento", "Chiamata"])
    for cr in crew:
        ws.append(list(cr))

    # Autofit
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col[0].column)
        ].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Public API
__all__ = [
    # Formato
    'Format',
    # Funzioni unificate
    'export_breakdown',
    'export_budget',
    'export_stripboard',
    'export_oneliner',
    'export_dood',
    'export_call_sheet',
    # Classi PDF (per uso avanzato)
    'BasePDFExporter',
    'OneLinerExporter',
    'DayOutOfDaysExporter',
    'CallSheetGenerator',
    'CallSheetPDF',
    'generate_call_sheet_pdf',
    # Flag disponibilità
    'HAS_FPDF',
]
